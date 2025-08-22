from ..serializers import employee_serializers
from ..serializers.hr_serializers import *
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from ..models import Employees, TrainingProgram
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from rest_framework.parsers import JSONParser
from django.shortcuts import get_object_or_404
from django.db.models import Q
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from django.contrib.auth import get_user_model
import csv
from datetime import datetime

# employee registration
@api_view(['POST'])
def register_employee(request):
    try:
        # Convert empty strings to None for optional fields
        data = request.data.copy()
        for field in ['bankName', 'bankAccount', 'pensionFund', 'nssaNumber', 
                     'zimraTaxNumber', 'payeNumber', 'aidsLevyNumber']:
            if field in data and data[field] == "":
                data[field] = None
        
        serializer = employee_serializers.EmployeeRegistrationSerializer(data=data)
        
        if serializer.is_valid():
            employee = serializer.save()
            return Response({
                "message": "Employee created successfully!",
                "employeeid": employee.employeeid
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
def get_all_employees(request): 
    users = Employees.objects.all()
    # serializer = employee_serializers.EmployeeRegistrationSerializer(users, many=True)
    serializer = employee_serializers.EmployeePayslipSerializer(users, many=True)

    
    return Response(serializer.data, status=status.HTTP_200_OK)



#This view handles the registration of training programs
#this function also handles the filtering of training programs based on mandatory status and category
# get all training programs
@csrf_exempt
@require_http_methods(["GET", "POST"])
def training_program_list(request):
    if request.method == 'GET':
        # Filtering capability
        mandatory = request.GET.get('mandatory')
        category = request.GET.get('category')
        
        programs = TrainingProgram.objects.all()
        
        if mandatory in ['true', 'false']:
            programs = programs.filter(mandatory=mandatory.lower() == 'true')
        if category:
            programs = programs.filter(category__iexact=category)
            
        serializer = TrainingProgramSerializer(programs, many=True)
        return JsonResponse({
            'count': programs.count(),
            'results': serializer.data
        }, safe=False)

    elif request.method == 'POST':
        try:
            data = JSONParser().parse(request)
            
            # Custom validation
            if not data.get('title'):
                return JsonResponse(
                    {'error': 'Title is required'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            serializer = TrainingProgramSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return JsonResponse(
                    serializer.data, 
                    status=status.HTTP_201_CREATED
                )
            return JsonResponse(
                serializer.errors, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return JsonResponse(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

#this function handles the retrieval, update, and deletion of a specific training program
# training program detail
@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def training_program_detail(request, pk):
    try:
        program = TrainingProgram.objects.get(pk=pk)
    except TrainingProgram.DoesNotExist:
        return JsonResponse(
            {'error': 'Training program not found'}, 
            status=status.HTTP_404_NOT_FOUND
        )

    if request.method == 'GET':
        serializer = TrainingProgramSerializer(program)
        return JsonResponse(serializer.data)

    elif request.method == 'PUT':
        try:
            data = JSONParser().parse(request)
            serializer = TrainingProgramSerializer(program, data=data)
            if serializer.is_valid():
                serializer.save()
                return JsonResponse(serializer.data)
            return JsonResponse(
                serializer.errors, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return JsonResponse(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    elif request.method == 'DELETE':
        try:
            program.delete()
            return JsonResponse(
                {'message': 'Training program deleted successfully'}, 
                status=status.HTTP_204_NO_CONTENT
            )
        except Exception as e:
            return JsonResponse(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



#
@api_view(['GET', 'POST'])
def training_session_list(request):
    """
    List all sessions or create a new session
    """
    if request.method == 'GET':
        sessions = TrainingSession.objects.all()
        serializer = TrainingSessionSerializer(sessions, many=True)
        return JsonResponse(serializer.data, safe=False)
    
    elif request.method == 'POST':
        data = JSONParser().parse(request)
        serializer = TrainingSessionSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=status.HTTP_201_CREATED)
        return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
def training_session_detail(request, pk):
    """
    Retrieve, update or delete a session
    """
    try:
        session = TrainingSession.objects.get(pk=pk)
    except TrainingSession.DoesNotExist:
        return JsonResponse(
            {'error': 'Training session not found'}, 
            status=status.HTTP_404_NOT_FOUND
        )

    if request.method == 'GET':
        serializer = TrainingSessionSerializer(session)
        return JsonResponse(serializer.data)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = TrainingSessionSerializer(session, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data)
        return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        session.delete()
        return JsonResponse(
            {'message': 'Training session deleted successfully'}, 
            status=status.HTTP_204_NO_CONTENT
        )
        
@api_view(['GET', 'POST'])
@csrf_exempt
def training_enrollment_list(request):
    if request.method == 'GET':
        enrollments = TrainingEnrollment.objects.all()
        serializer = TrainingEnrollmentSerializer(enrollments, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        print("Received data:", request.data)  # Debugging line
        serializer = TrainingEnrollmentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def training_enrollment_detail(request, pk):
    try:
        enrollment = TrainingEnrollment.objects.get(pk=pk)
    except TrainingEnrollment.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = TrainingEnrollmentSerializer(enrollment)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = TrainingEnrollmentSerializer(enrollment, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        enrollment.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
@api_view(['GET', 'POST'])
def certification_list(request):
    if request.method == 'GET':
        certifications = TrainingCertification.objects.all()
        serializer = CertificationSerializer(certifications, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = CertificationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
def certification_detail(request, pk):
    certification = get_object_or_404(TrainingCertification, pk=pk)
    
    if request.method == 'GET':
        serializer = CertificationSerializer(certification)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = CertificationSerializer(certification, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        certification.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['GET'])
def certification_search(request):
    search_term = request.query_params.get('search', '')
    certifications = TrainingCertification.objects.filter(
        Q(employee__icontains=search_term) | 
        Q(program__icontains=search_term))
    serializer = CertificationSerializer(certifications, many=True)
    return Response(serializer.data)

  
# ---------------- Attendance Endpoints -----------------
@api_view(['GET', 'POST'])
def attendance_list(request):
    """
    GET: List all attendance records formatted for the frontend.
    POST: Create a new attendance record (raw model fields).
    """
    if request.method == 'GET':
        qs = AttendanceRecord.objects.select_related('employee').all().order_by('-date', 'employee__first_name')
        serializer = AttendanceRecordAPISerializer(qs, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = AttendanceRecordSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
def attendance_delete(request, pk: int):
    """Delete a single attendance record by id."""
    try:
        rec = AttendanceRecord.objects.get(pk=pk)
    except AttendanceRecord.DoesNotExist:
        return Response({'detail': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    rec.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def attendance_bulk_upload(request):
    """
    Bulk upload attendance records from a CSV or Excel file.
    Expected columns (case-insensitive; spaces/underscores ignored):
      - employeeId or job_number or jobno
      - date (YYYY-MM-DD or common date formats)
      - loginTime (HH:MM or HH:MM:SS)
      - logoutTime (HH:MM or HH:MM:SS)

    Tries to resolve employee via CustomUser fields in order:
      employeeid == job_number, then username == job_number, then email == job_number.
    Rows where employee cannot be resolved are skipped with an error.
    """
    upload = request.FILES.get('file')
    if not upload:
        return Response({'detail': 'No file provided under "file" key.'}, status=status.HTTP_400_BAD_REQUEST)

    filename = upload.name.lower()
    summary = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}

    def norm(s: str):
        return (s or '').strip().lower().replace(' ', '').replace('_', '')

    def parse_date(val):
        if not val:
            return None
        if isinstance(val, datetime):
            return val.date()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y']:
            try:
                return datetime.strptime(str(val).strip(), fmt).date()
            except ValueError:
                continue
        return None

    def parse_time(val):
        if not val:
            return None
        if isinstance(val, datetime):
            return val.time()
        for fmt in ['%H:%M:%S', '%H:%M']:
            try:
                return datetime.strptime(str(val).strip(), fmt).time()
            except ValueError:
                continue
        return None

    def resolve_user(job_number: str):
        if not job_number:
            return None
        User = get_user_model()
        # Try match by custom employeeid first, then username, then email
        user = User.objects.filter(employeeid=job_number).first()
        if user:
            return user
        user = User.objects.filter(username=job_number).first()
        if user:
            return user
        user = User.objects.filter(email=job_number).first()
        return user

    try:
        # CSV handling
        if filename.endswith('.csv'):
            data = upload.read().decode('utf-8', errors='ignore')
            reader = csv.DictReader(data.splitlines())
            headers = [norm(h) for h in reader.fieldnames or []]

            # Map dynamic header names
            def get_val(row, keys):
                for k in keys:
                    for h in row.keys():
                        if norm(h) == norm(k):
                            return row.get(h)
                return None

            for idx, row in enumerate(reader, start=2):  # start=2 accounts for header row
                job_number = get_val(row, ['employeeId', 'job_number', 'jobno', 'job no'])
                date_val = get_val(row, ['date'])
                time_in_val = get_val(row, ['loginTime', 'time_in', 'time in'])
                time_out_val = get_val(row, ['logoutTime', 'time_out', 'time out'])

                day = parse_date(date_val)
                t_in = parse_time(time_in_val)
                t_out = parse_time(time_out_val)
                user = resolve_user(job_number)

                if not user:
                    summary['skipped'] += 1
                    summary['errors'].append({'row': idx, 'error': f'User not found for job number: {job_number}'})
                    continue
                if not day:
                    summary['skipped'] += 1
                    summary['errors'].append({'row': idx, 'error': 'Invalid or missing date'})
                    continue

                obj, created = AttendanceRecord.objects.update_or_create(
                    employee=user,
                    date=day,
                    defaults={'time_in': t_in, 'time_out': t_out, 'job_number': job_number or ''}
                )
                summary['created' if created else 'updated'] += 1

        # XLSX handling via openpyxl
        elif filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            try:
                from openpyxl import load_workbook
            except Exception:
                return Response({'detail': 'openpyxl is required to parse Excel files.'}, status=status.HTTP_400_BAD_REQUEST)

            wb = load_workbook(upload, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return Response({'detail': 'Empty Excel file.'}, status=status.HTTP_400_BAD_REQUEST)
            header = [str(h) if h is not None else '' for h in rows[0]]

            def idx_for(keys):
                normed = [norm(h) for h in header]
                for key in keys:
                    if norm(key) in normed:
                        return normed.index(norm(key))
                return -1

            i_job = idx_for(['employeeId', 'job_number', 'jobno', 'job no'])
            i_date = idx_for(['date'])
            i_in = idx_for(['loginTime', 'time_in', 'time in'])
            i_out = idx_for(['logoutTime', 'time_out', 'time out'])

            for r_idx, r in enumerate(rows[1:], start=2):
                job_number = r[i_job] if i_job >= 0 and i_job < len(r) else None
                date_val = r[i_date] if i_date >= 0 and i_date < len(r) else None
                time_in_val = r[i_in] if i_in >= 0 and i_in < len(r) else None
                time_out_val = r[i_out] if i_out >= 0 and i_out < len(r) else None

                day = parse_date(date_val)
                t_in = parse_time(time_in_val)
                t_out = parse_time(time_out_val)
                user = resolve_user(str(job_number).strip() if job_number is not None else None)

                if not user:
                    summary['skipped'] += 1
                    summary['errors'].append({'row': r_idx, 'error': f'User not found for job number: {job_number}'})
                    continue
                if not day:
                    summary['skipped'] += 1
                    summary['errors'].append({'row': r_idx, 'error': 'Invalid or missing date'})
                    continue

                obj, created = AttendanceRecord.objects.update_or_create(
                    employee=user,
                    date=day,
                    defaults={'time_in': t_in, 'time_out': t_out, 'job_number': str(job_number) if job_number else ''}
                )
                summary['created' if created else 'updated'] += 1
        else:
            return Response({'detail': 'Unsupported file type. Please upload a .csv or .xlsx file.'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'message': 'Upload processed', **summary}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


