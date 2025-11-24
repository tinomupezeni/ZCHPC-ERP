# hr/services.py
import csv
from datetime import datetime
from typing import Dict, Any

from openpyxl import load_workbook
from ..hr_models import (
    Employees, 
    Department, 
    TrainingProgram, 
    TrainingSession,
    TrainingEnrollment,
    TrainingCertification,
    AttendanceRecord,
)

# --- Employees Services ---

def create_employee(data: Dict[str, Any]) -> Employees:
    """
    Creates a new Employees instance.
    Business logic (like creating a user account or sending
    a welcome email) would go here.
    """
    # TODO: Add logic to create/link a CustomUser if one isn't provided
    employee = Employees.objects.create(**data)
    return employee

def update_employee(instance: Employees, data: Dict[str, Any]) -> Employees:
    """
    Updates an existing Employees instance.
    """
    for attr, value in data.items():
        setattr(instance, attr, value)
    instance.save()
    return instance

def deactivate_employee(instance: Employees):
    """
    Soft-deletes an employee by setting them to inactive
    instead of permanently deleting.
    """
    instance.is_active = False
    instance.save()

# --- Department Services ---

def create_department(data: Dict[str, Any]) -> Department:
    """
    Creates a new Department.
    """
    department = Department.objects.create(**data)
    return department

def update_department(instance: Department, data: Dict[str, Any]) -> Department:
    """
    Updates an existing Department.
    """
    for attr, value in data.items():
        setattr(instance, attr, value)
    instance.save()
    return instance

def delete_department(instance: Department):
    """
    Deletes a Department.
    """
    # TODO: Add logic to check if department has employees before deleting
    # if instance.employees.count() > 0:
    #    raise ValidationError("Cannot delete department with active employees.")
    instance.delete()


# --- Training Program Services ---

def create_training_program(data: Dict[str, Any]) -> TrainingProgram:
    """
    Creates a new TrainingProgram.
    """
    program = TrainingProgram.objects.create(**data)
    return program

def update_training_program(instance: TrainingProgram, data: Dict[str, Any]) -> TrainingProgram:
    """
    Updates an existing TrainingProgram.
    """
    for attr, value in data.items():
        setattr(instance, attr, value)
    instance.save()
    return instance

def delete_training_program(instance: TrainingProgram):
    """
    Deletes a TrainingProgram.
    """
    # TODO: Add logic to check for active sessions
    instance.delete()


# --- Training Session Services ---

def create_training_session(data: Dict[str, Any]) -> TrainingSession:
    """
    Creates a new TrainingSession.
    """
    session = TrainingSession.objects.create(**data)
    return session

def update_training_session(instance: TrainingSession, data: Dict[str, Any]) -> TrainingSession:
    """
    Updates an existing TrainingSession.
    """
    for attr, value in data.items():
        setattr(instance, attr, value)
    instance.save()
    return instance

def delete_training_session(instance: TrainingSession):
    """
    Deletes a TrainingSession.
    """
    # TODO: Add logic to check for enrollments
    instance.delete()
    
def create_enrollment(data: Dict[str, Any]) -> TrainingEnrollment:
    enrollment = TrainingEnrollment.objects.create(**data)
    return enrollment

def update_enrollment(instance: TrainingEnrollment, data: Dict[str, Any]) -> TrainingEnrollment:
    for attr, value in data.items():
        setattr(instance, attr, value)
    instance.save()
    return instance

def delete_enrollment(instance: TrainingEnrollment):
    instance.delete()

# --- NEW: Training Certification Services ---

def create_certification(data: Dict[str, Any]) -> TrainingCertification:
    certification = TrainingCertification.objects.create(**data)
    return certification

def update_certification(instance: TrainingCertification, data: Dict[str, Any]) -> TrainingCertification:
    for attr, value in data.items():
        setattr(instance, attr, value)
    instance.save()
    return instance

def delete_certification(instance: TrainingCertification):
    instance.delete()

# --- NEW: Attendance Record Services ---

def create_attendance_record(data: Dict[str, Any]) -> AttendanceRecord:
    record = AttendanceRecord.objects.create(**data)
    return record

def delete_attendance_record(instance: AttendanceRecord):
    instance.delete()

# --- NEW: Attendance Bulk Upload Service ---
# (This is the complex logic from your attendance_bulk_upload view)

def _norm_header(s: str):
    return (s or '').strip().lower().replace(' ', '').replace('_', '')

def _parse_date(val):
    if not val: return None
    if isinstance(val, datetime): return val.date()
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y']:
        try: return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError: continue
    return None

def _parse_time(val):
    if not val: return None
    if isinstance(val, datetime): return val.time()
    for fmt in ['%H:%M:%S', '%H:%M']:
        try: return datetime.strptime(str(val).strip(), fmt).time()
        except ValueError: continue
    return None

def _resolve_employee(emp_id_str: str):
    if not emp_id_str:
        return None
    # Use our new Employees model and its employee_id field
    return Employees.objects.filter(employee_id=emp_id_str).first()

def process_attendance_upload(file_obj) -> Dict[str, Any]:
    """
    Processes an uploaded CSV or XLSX file for bulk attendance.
    """
    filename = file_obj.name.lower()
    summary = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
    
    records_to_process = []
    
    try:
        # --- CSV HANDLING ---
        if filename.endswith('.csv'):
            data = file_obj.read().decode('utf-8', errors='ignore')
            reader = csv.DictReader(data.splitlines())
            
            def get_val(row, keys):
                for k in keys:
                    for h in row.keys():
                        if _norm_header(h) == _norm_header(k):
                            return row.get(h)
                return None
            
            for idx, row in enumerate(reader, start=2):
                records_to_process.append({
                    'row_num': idx,
                    'emp_id': get_val(row, ['employeeId', 'employee_id', 'job_number']),
                    'date_val': get_val(row, ['date']),
                    'time_in_val': get_val(row, ['loginTime', 'time_in', 'time in']),
                    'time_out_val': get_val(row, ['logoutTime', 'time_out', 'time out']),
                })
        
        # --- XLSX HANDLING ---
        elif filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            wb = load_workbook(file_obj, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise ValueError("Empty Excel file.")
            
            header = [str(h) if h is not None else '' for h in rows[0]]
            normed_header = [_norm_header(h) for h in header]
            
            def idx_for(keys):
                for key in keys:
                    if _norm_header(key) in normed_header:
                        return normed_header.index(_norm_header(key))
                return -1

            i_job = idx_for(['employeeId', 'employee_id', 'job_number'])
            i_date = idx_for(['date'])
            i_in = idx_for(['loginTime', 'time_in', 'time in'])
            i_out = idx_for(['logoutTime', 'time_out', 'time out'])

            for r_idx, r in enumerate(rows[1:], start=2):
                records_to_process.append({
                    'row_num': r_idx,
                    'emp_id': r[i_job] if i_job != -1 and i_job < len(r) else None,
                    'date_val': r[i_date] if i_date != -1 and i_date < len(r) else None,
                    'time_in_val': r[i_in] if i_in != -1 and i_in < len(r) else None,
                    'time_out_val': r[i_out] if i_out != -1 and i_out < len(r) else None,
                })
        else:
            raise ValueError("Unsupported file type. Please upload a .csv or .xlsx file.")

        # --- PROCESS ALL RECORDS ---
        for record in records_to_process:
            emp_id_str = str(record['emp_id']).strip() if record['emp_id'] else None
            employee = _resolve_employee(emp_id_str)
            day = _parse_date(record['date_val'])
            t_in = _parse_time(record['time_in_val'])
            t_out = _parse_time(record['time_out_val'])

            if not employee:
                summary['skipped'] += 1
                summary['errors'].append({'row': record['row_num'], 'error': f'Employees not found for ID: {emp_id_str}'})
                continue
            if not day:
                summary['skipped'] += 1
                summary['errors'].append({'row': record['row_num'], 'error': 'Invalid or missing date'})
                continue
            
            obj, created = AttendanceRecord.objects.update_or_create(
                employee=employee,
                date=day,
                defaults={'time_in': t_in, 'time_out': t_out}
            )
            summary['created' if created else 'updated'] += 1

    except Exception as e:
        summary['errors'].append({'row': 'File', 'error': f'General processing error: {str(e)}'})

    return summary
